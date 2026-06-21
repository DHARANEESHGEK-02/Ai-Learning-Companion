"""
Report Generator — produces PDF and Excel evaluation reports.
PDF via ReportLab, Excel via openpyxl.
"""
import io
from datetime import datetime
from typing import List, Dict, Any


def generate_pdf_report(data: dict) -> bytes:
    """Generate a comprehensive PDF evaluation report."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                             topMargin=2*cm, bottomMargin=2*cm,
                             leftMargin=2*cm, rightMargin=2*cm)

    styles = getSampleStyleSheet()
    primary = HexColor("#1E3A5F")
    accent = HexColor("#2196F3")
    light_bg = HexColor("#F0F4F8")
    success = HexColor("#4CAF50")
    warning = HexColor("#FF9800")
    danger = HexColor("#F44336")

    title_style = ParagraphStyle("Title", parent=styles["Title"],
                                  textColor=primary, fontSize=20, spaceAfter=6)
    h2_style = ParagraphStyle("H2", parent=styles["Heading2"],
                               textColor=primary, fontSize=13, spaceBefore=12, spaceAfter=6)
    body_style = ParagraphStyle("Body", parent=styles["Normal"], fontSize=10, leading=14)
    small_style = ParagraphStyle("Small", parent=styles["Normal"], fontSize=9, textColor=HexColor("#555555"))

    elements = []
    elements.append(Paragraph("AI Learning Companion", title_style))
    elements.append(Paragraph("Evaluation Report", ParagraphStyle("Sub", parent=styles["Heading3"],
                                                                    textColor=accent, fontSize=14, spaceAfter=4)))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}", small_style))
    elements.append(HRFlowable(width="100%", thickness=2, color=primary))
    elements.append(Spacer(1, 0.4*cm))

    student_name = data.get("student_name", "N/A")
    exam_title = data.get("exam_title", "N/A")
    total_score = data.get("total_score", 0)
    max_score = data.get("max_score", 100)
    percentage = round(total_score / max_score * 100, 1) if max_score else 0
    plagiarism = data.get("plagiarism_score", 0) or 0

    grade = _get_grade(percentage)
    score_color = success if percentage >= 75 else (warning if percentage >= 50 else danger)

    summary_data = [
        ["Student", student_name, "Exam", exam_title],
        ["Score", f"{total_score} / {max_score}", "Percentage", f"{percentage}%"],
        ["Grade", grade, "Plagiarism Risk", f"{round(plagiarism * 100, 1)}%"],
    ]
    summary_table = Table(summary_data, colWidths=[3.5*cm, 6*cm, 3.5*cm, 4*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), light_bg),
        ('BACKGROUND', (0, 0), (0, -1), primary),
        ('BACKGROUND', (2, 0), (2, -1), primary),
        ('TEXTCOLOR', (0, 0), (0, -1), white),
        ('TEXTCOLOR', (2, 0), (2, -1), white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('PADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor("#CCCCCC")),
        ('ROUNDEDCORNERS', [4]),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.5*cm))

    evaluations = data.get("evaluations", [])
    if evaluations:
        elements.append(Paragraph("Question-wise Evaluation", h2_style))
        for ev in evaluations:
            q_num = ev.get("question_number", "?")
            q_text = ev.get("question_text", "")[:150]
            score = ev.get("score", 0)
            q_max = ev.get("max_score", 0)
            feedback = ev.get("feedback", "")
            missing = ev.get("missing_concepts", [])
            strengths = ev.get("strengths", [])
            sim = ev.get("similarity_score", 0)

            q_pct = round(score / q_max * 100) if q_max else 0
            q_color = success if q_pct >= 75 else (warning if q_pct >= 50 else danger)

            q_header = [[
                Paragraph(f"Q{q_num}. {q_text}", body_style),
                Paragraph(f"<b>{score}/{q_max}</b> ({q_pct}%)", ParagraphStyle("Score",
                    parent=styles["Normal"], fontSize=11, textColor=q_color, alignment=2))
            ]]
            q_table = Table(q_header, colWidths=[13*cm, 4*cm])
            q_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), light_bg),
                ('PADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, HexColor("#CCCCCC")),
            ]))
            elements.append(q_table)

            if feedback:
                elements.append(Paragraph(f"<i>Feedback:</i> {feedback}", small_style))
            if strengths:
                elements.append(Paragraph(f"<b>Strengths:</b> {', '.join(strengths)}", small_style))
            if missing:
                elements.append(Paragraph(f"<b>Missing Concepts:</b> {', '.join(missing)}", small_style))
            elements.append(Spacer(1, 0.3*cm))

    learning_plan = data.get("learning_plan")
    if learning_plan:
        elements.append(Paragraph("Personalized Learning Plan", h2_style))
        recs = learning_plan.get("recommendations", [])
        for i, rec in enumerate(recs, 1):
            elements.append(Paragraph(f"{i}. {rec}", body_style))
        elements.append(Spacer(1, 0.3*cm))

    doc.build(elements)
    return buffer.getvalue()


def generate_excel_report(submissions: List[Dict[str, Any]]) -> bytes:
    """Generate an Excel batch evaluation report."""
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation Report"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill = PatternFill("solid", fgColor="F0F4F8")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(["AI Learning Companion — Batch Evaluation Report"])
    ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws.append([f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}"])
    ws.append([])

    headers = [
        "Student Name", "Email", "Exam", "Score", "Max Score",
        "Percentage (%)", "Grade", "Plagiarism (%)", "Status", "Submitted At"
    ]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for i, sub in enumerate(submissions, 5):
        score = sub.get("total_score") or 0
        max_s = sub.get("max_score") or 100
        pct = round(score / max_s * 100, 1) if max_s else 0
        grade = _get_grade(pct)
        plag = round((sub.get("plagiarism_score") or 0) * 100, 1)

        row = [
            sub.get("student_name", ""),
            sub.get("email", ""),
            sub.get("exam_title", ""),
            score,
            max_s,
            pct,
            grade,
            plag,
            sub.get("status", ""),
            sub.get("submitted_at", ""),
        ]
        ws.append(row)

        fill = PatternFill("solid", fgColor="FFFFFF") if i % 2 == 0 else alt_fill
        score_cell = ws.cell(row=i, column=6)
        if pct >= 75:
            score_cell.font = Font(color="4CAF50", bold=True)
        elif pct >= 50:
            score_cell.font = Font(color="FF9800", bold=True)
        else:
            score_cell.font = Font(color="F44336", bold=True)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    col_widths = [20, 25, 25, 8, 10, 14, 8, 14, 12, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Summary Statistics"])
    ws2["A1"].font = Font(bold=True, size=13, color="1E3A5F")
    ws2.append([])
    evaluated = [s for s in submissions if s.get("total_score") is not None]
    scores = [s["total_score"] / max(1, s.get("max_score") or 1) * 100 for s in evaluated]
    ws2.append(["Total Submissions", len(submissions)])
    ws2.append(["Evaluated", len(evaluated)])
    ws2.append(["Average Score (%)", round(sum(scores) / len(scores), 1) if scores else 0])
    ws2.append(["Highest Score (%)", round(max(scores), 1) if scores else 0])
    ws2.append(["Lowest Score (%)", round(min(scores), 1) if scores else 0])
    ws2.append(["Pass Rate (>=50%)", f"{round(len([s for s in scores if s >= 50]) / max(1, len(scores)) * 100, 1)}%"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _get_grade(percentage: float) -> str:
    if percentage >= 90:
        return "O (Outstanding)"
    elif percentage >= 75:
        return "A+ (Excellent)"
    elif percentage >= 60:
        return "A (Very Good)"
    elif percentage >= 50:
        return "B (Good)"
    elif percentage >= 40:
        return "C (Pass)"
    else:
        return "F (Fail)"
